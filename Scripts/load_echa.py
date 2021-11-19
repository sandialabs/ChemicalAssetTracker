from openpyxl import Workbook
from openpyxl import load_workbook
from utils import *
from cmsdb import *
import re
import datetime


CAS_REGEX =  re.compile('(\\d+-\\d+-\\d+)')

HAZARD_CLASSES = [

# misspellings
"Muta. 1B A ",
"Carc. 2.",

"Unst. Expl.", "Expl. 1.1", "Expl. 1.2", "Expl. 1.3", "Expl. 1.4", "Expl. 1.5", "Expl. 1.6",
"Flam. Gas 1", "Flam. Gas 1A", "Flam. Gas 1B", "Flam. Gas 2", "Pyr. Gas", "Chem. Unst. Gas A", "Chem. Unst. Gas B",
"Aerosol 1", "Aerosol 2", "Aerosol 3",
"Ox. Gas 1",
"Pres. Gas",
"Flam. Liq. 1", "Flam. Liq. 2", "Flam. Liq. 3",
"Flam. Sol. 1", "Flam. Sol. 2",
"Self-react. A", "Self-react. B", "Self-react. C", "Self-react. D", "Self-react. E", "Self-react. F", "Self-react. G",
"Pyr. Liq. 1",
"Pyr. Sol. 1",
"Self-heat. 1", "Self-heat. 2",
"Water-react. 1",  "Water-react. 2", "Water-react. 3",
"Ox. Liq. 1", "Ox. Liq. 2", "Ox. Liq. 3",
"Ox. Sol. 1", "Ox. Sol. 2", "Ox. Sol. 3",
"Org. Perox. A", "Org. Perox. B", "Org. Perox. C", "Org. Perox. D", "Org. Perox. E", "Org. Perox. F", "Org. Perox. G",
"Met. Corr. 1",
"Desen. Expl. 1", "Desen. Expl. 2", "Desen. Expl. 3", "Desen. Expl. 4",
"Acute Tox. 1", "Acute Tox. 2", "Acute Tox. 3", "Acute Tox. 4",
"Skin Corr. 1A", "Skin Corr. 1B", "Skin Corr. 1C", "Skin Corr. 1", "Skin Irrit. 2",
"Eye Dam. 1", "Eye Irrit. 2",
"Resp. Sens. 1A", "Resp. Sens. 1B", "Resp. Sens. 1",
"Skin Sens. 1A", "Skin Sens. 1B", "Skin Sens. 1",
"Muta. 1A", "Muta. 1B", "Muta. 2",
"Carc. 1A", "Carc. 1B", "Carc. 2",
"Repr. 1A", "Repr. 1B", "Repr. 2", "Lact.",
"STOT SE 1", "STOT SE 2", "STOT SE 3",
"STOT RE 1", "STOT RE 2",
"Asp. Tox. 1",
"Aquatic Acute 1", "Aquatic Chronic 1", "Aquatic Chronic 2", "Aquatic Chronic 3", "Aquatic Chronic 4",
"Ozone 1",

# misspellings
"Press. Gas",
"Asp Tox. 1",
"Unst. Expl",
"Self-heat 1",
"Carc. 1Β",
"Expl.",
"Eye Dam.1"


]





def has_value(text):
    if text == None:
        return False
    text = str(text)
    if text and len(text.strip()) > 0:
        return True
    else:
        return False


def extract_cas_number(text):
    m = CAS_REGEX.match(text)
    if m:
        return m[1]
    else:
        return None



def split_string_values(text, dlm=' ', trim_whitespace=True):
    result = []
    if text:
        lines = text.split('|')
        for line in lines:
            parts = line.split(dlm)
            if trim_whitespace:
                for x in parts:
                    trimmed = x.strip()
                    if len(trimmed) > 0:
                        result.append(trimmed)
            else:
                for x in parts:
                    result.append(x)
    return result


def clean_hazard_class_and_category_codes(text):
    if (text):
        return text.replace("'", '').replace('Gas.', 'Gas').replace('\n', ' ').replace('|', ' ').replace('*', '').replace('    ', ' ').replace('   ', ' ').replace('  ', ' ').strip()
    else:
        return text

def find_hazard_class_and_category_code(text):
    lowtext = text.lower()
    for prefix in HAZARD_CLASSES:
        if lowtext.startswith(prefix.lower()):
            l = len(prefix)
            return(prefix, text[l:].strip())
    return (None, None)


def parse_hazard_class_and_category_codes(text):
    result = []
    if text:
        text = clean_hazard_class_and_category_codes(text)
        while len(text) > 0:
            (prefix, suffix) = find_hazard_class_and_category_code(text)
            if prefix == None:
                raise Exception("Unrecognized hazard category code: " + text)
            result.append(prefix)
            text = suffix
    return result

def check_hazard_class_and_category_codes(hazards):
    result = []
    for hazard in hazards:
        text = clean_hazard_class_and_category_codes(hazard.m_hazard_class_and_category_codes)
        while len(text) > 0:
            (prefix, suffix) = find_hazard_class_and_category_code(text)
            if prefix == None:
                if text not in result:
                    result.append(text)
                    print(f"{hazard.m_rownum}: {hazard.m_hazard_class_and_category_codes}")
                text = ''
            else:
                text = suffix
    for x in result:
        print(x)


class HazardInfo:

    def __init__(self, row):
        self.m_rownum = row.m_rownum
        self.m_datacells = row.nonempty_cells()
        self.m_index_number = None
        self.m_chemical = None
        self.m_ecnumber = None
        self.m_casnumbers = None
        self.m_hazard_class_and_category_codes = None
        self.m_class_hazard_statement_codes = None
        self.m_pictogram_codes = None
        self.m_labeling_hazard_statement_codes = None
        self.m_supplemental_codes = None
        self.Valid = False

        if len(self.m_datacells) > 0:
            idcellnum = self.m_datacells[0].m_colnum
            if idcellnum == 4:
                # type 1: 
                #   Cell 4:  Index number
                #   Cell 12: chemical
                #   Cell 31: ECNumber
                #   Cell 42: CASNumber
                #   Cell 50: Classification: Hazard class and category codes
                #   Cell 60: Classification: Hazard statement code(s)
                #   Cell 69: Labeling: Pictogram, signal word code(s)
                #   Cell 76: Labeling: Hazard Sttement code(s)
                #   Cell 80: Supplemental hazard statement code(s)
                self.m_index_number = self.find_cell_value(4)
                self.m_chemical = self.find_cell_value(12)
                self.m_ecnumber = self.find_cell_value(31)
                self.m_casnumbers = self.find_cell_value(42)
                self.m_hazard_class_and_category_codes = self.find_cell_value(50)
                self.m_class_hazard_statement_codes = self.find_cell_value(60)
                self.m_pictogram_codes = self.find_cell_value(69)
                self.m_labeling_hazard_statement_codes = self.find_cell_value(76)
                self.m_supplemental_codes = self.find_cell_value(80)

            if idcellnum == 0:
                # type 2: ECNumber in cell 26, CASNumber in cell 36
                #   Cell 0:  Index number
                #   Cell 7: chemical
                #   Cell 26: ECNumber
                #   Cell 36: CASNumber
                #   Cell 45: Classification: Hazard class and category codes
                #   Cell 56: Classification: Hazard statement code(s)
                #   Cell 65: Labeling: Pictogram, signal word code(s)
                #   Cell 72: Labeling: Hazard Sttement code(s)
                #   Cell 78: Supplemental hazard statement code(s)
                self.m_index_number = self.find_cell_value(0)
                self.m_chemical = self.find_cell_value(7)
                self.m_ecnumber = self.find_cell_value(26)
                self.m_casnumbers = self.find_cell_value(36)
                self.m_hazard_class_and_category_codes = self.find_cell_value(45)
                self.m_class_hazard_statement_codes = self.find_cell_value(56)
                self.m_pictogram_codes = self.find_cell_value(65)
                self.m_labeling_hazard_statement_codes = self.find_cell_value(72)
                self.m_supplemental_codes = self.find_cell_value(78)
            
            self.Valid = has_value(self.m_casnumbers) and self.m_casnumbers != "None" and has_value(self.m_chemical) > 0 and has_value(self.m_pictogram_codes) > 0 and has_value(self.m_labeling_hazard_statement_codes) > 0


    def find_cell(self, colnum):
        cols = self.m_datacells
        match = [col for col in cols if col.m_colnum == colnum]
        if len(match) == 1:
            return match[0]
        else:
            return None

    def find_cell_value(self, colnum):
        result = None
        col = self.find_cell(colnum)
        if col:
            result = col.nice_value()
        return result


    def get_cas_numbers(self):
        parts = split_string_values(self.m_casnumbers)
        result = []
        for part in parts:
            casnum = extract_cas_number(part)
            if casnum:
                result.append(casnum)
        return result

    def get_pictograms(self):
        result = []
        parts = split_string_values(self.m_pictogram_codes)
        for part in parts:
            if part != 'Dgr':
                result.append(part)
        return result


    def get_hazard_codes(self):
        result = []
        parts = split_string_values(self.m_class_hazard_statement_codes)
        for part in parts:
            if part[0:1] == 'H':
                result.append(part)
        return result

    def get_hazard_class_and_category_codes(self):
        return parse_hazard_class_and_category_codes(self.m_hazard_class_and_category_codes)

    def show(self):
        print(f"Hazard Information - row {self.m_rownum}")
        print(f"    Index #:                {self.m_index_number}")
        print(f"    Chemical:               {self.m_chemical}")
        print(f"    EC #:                   {self.m_ecnumber}")
        print(f'    CAS String              "{self.m_casnumbers}"')
        print(f"    CAS Numbers:            {self.get_cas_numbers()}")
        print(f'    Hazard Class/Cat Str:   "{self.m_hazard_class_and_category_codes}"')
        print(f'    Hazard Class/Cat Codes: {self.get_hazard_class_and_category_codes()}')
        print(f'    Hazard Code String:     "{self.m_class_hazard_statement_codes}"')
        print(f"    Hazard Code(s)          {self.get_hazard_codes()}")
        print(f'    Pictogram String:       "{self.m_pictogram_codes}"')
        print(f"    Pictograms:             {self.get_pictograms()}")
        print(f"    Label Stmt Code(s):     {self.m_labeling_hazard_statement_codes}")
        print(f"    Supplemental Code(s):   {self.m_supplemental_codes}")

    def in_database(self, db):
        missing_count = 0
        casnumbers = self.get_cas_numbers();
        for casnumber in casnumbers:
            if not db.row_exists(f"select * from HazardCodes where CASNumber = '{casnumber}'"):
                print(f"CAS # {casnumber} is not in the HazardCodes table")
                missing_count += 1
        return missing_count

    def update_hazard_codes(self, db, casnumber):
        casnumbers = self.get_cas_numbers()
        pictograms = self.get_pictograms()
        hazard_codes = self.get_hazard_codes()
        hazard_class_codes = self.get_hazard_class_and_category_codes()

        for casnumber in casnumbers:
            casnumber = casnumber.strip()
            if not db.row_exists(f"select * from HazardCodes where CASNumber = '{casnumber}'"):
                for ix, code in enumerate(hazardcodes):
                    hclass = hazard_class_codes[ix]
                    if len(code) > 1:
                        db.execute_nonquery("insert into HazardCodes (GHSCode, CASNumber, HazardClass) values (%s, %s, %s)", [code.strip(), casnumber, hclass], commit=False)
                db.commit()


class Cell:

    def __init__(self, pxl_cell, rownum, colnum):

        self.m_cell = pxl_cell
        self.m_rownum = rownum
        self.m_colnum = colnum
        self.m_cas_regex = re.compile('(\\d+-\\d+-\\d+)')

    def __str__(self):
        return f"<Cell {self.m_rownum},{self.m_colnum}>"

    def __repr__(self):
        return f"<Cell {self.m_rownum},{self.m_colnum}>"

    def value(self):
        return self.m_cell.value

    def nice_value(self):
        val = self.value()
        if isinstance(val, str):
            val = val.replace('\n', "|")
        elif isinstance(val, datetime.datetime):
            # big kludge
            # pyopenxl seems to be converting cell values like '7637-07-02' into datetime.dateetime objects
            val = f"{val}"
            if ' ' in val:
                sppos = val.index(' ')
                val = val[0:sppos]
            print(f"Datetime found in cell {self.m_rownum},{self.m_colnum}: {val}")
            #print(f'Date value found at {self.m_rownum},{self.m_colnum}: "{val}"')
        return val

    def has_value(self):
        return self.m_cell.value != None

    def has_cas_number(self):
        result = False
        value = self.value()
        if isinstance(value, str):
            m = self.m_cas_regex.match(value)
            if m:
                result = True
        return result


class Row:

    def __init__(self, pxl_row, rownum):
        self.m_row = pxl_row
        self.m_rownum = rownum
        self.m_cellcount = len(pxl_row)
        self.m_cursor = 0

    def __getitem__(self, n):
        if n < self.m_cellcount:
            return Cell(self.m_row[n], self.m_rownum, n)
        else:
            return None

    def __str__(self):
        return f"<Row {self.m_rownum}>"

    def __repr__(self):
        return f"<Row {self.m_rownum}>"

    def __iter__(self):
        self.m_cursor = 0
        return self
    
    def __next__(self):
        if self.m_cursor < self.m_cellcount:
            result = self[self.m_cursor]
            self.m_cursor += 1
            return result
        else:
            raise StopIteration

    def cells(self):
        result = []
        for i in range(self.m_cellcount):
            result.append(Cell(self.m_row[i], self.m_rownum, i))
        return result

    def nonempty_cells(self):
        return [cell for cell in self.cells() if cell.has_value()]

    
    def get_id_cells(self):
        matches = [cell for cell in self.nonempty_cells() if cell.has_cas_number()]
        return matches

    def has_hazard_data(self):
        data = self.get_id_cells()
        return len(data) > 1

    def show(self):
        print(f"Row {self.m_rownum}")
        cells = self.nonempty_cells()
        if len(cells) > 0:
            for cell in cells:
                val = cell.value()
                if isinstance(val, str):
                    val = val.replace('\n', '\\n')
                print(f"    {cell.m_colnum}: {val}")

    def value(self, n):
        cell = self[n]
        return cell.value()

class Sheet:

    def __init__(self, pxl_worksheet):
        self.m_worksheet = pxl_worksheet
        self.m_cursor = 0
        self.m_rows = self.all_rows()
        self.m_hazards = []
        self.m_rowcount = len(self.m_rows)


    def __getitem__(self, n):
        if n <= self.m_rowcount:
            return self.m_rows[n]

    def __iter__(self):
        self.m_cursor = 0
        return self
    
    def __next__(self):
        if self.m_cursor < self.m_rowcount:
            result = self.m_rows[m_cursor]
            self.m_cursor += 1
            return result
        else:
            raise StopIteration

    def all_rows(self):
        result = []
        for (i, row) in enumerate(self.m_worksheet):
            result.append(Row(row, i+1))
        return result

    
    def get_hazard_data(self, verbose = False):
        if len(self.m_hazards) > 0:
            return self.m_hazards
        if verbose:
            print("Collecting hazard data ...")
        result = []
        rows = self.m_rows
        rowcount = len(rows)
        valid_count = 0
        current_row = 0
        try:
            for (i, row) in enumerate(rows):
                current_row = i
                if row.has_hazard_data():
                    info = HazardInfo(row)
                    if info.Valid:
                        result.append(info)
                        valid_count += 1
            if verbose:
                print("")
                print(f"Found {len(result)} valid data rows in {rowcount} spread rows")
        except:
            print(f"Exception processing row {current_row}")
            print(" ")
            raise
        self.m_hazards = result
        return result

    def show_id_rows(self):
        for i in range(self.m_rowcount):
            row = self[i]
            data = row.get_id_cells()
            if (len(data) > 0):
                print(f"{row}")
                for cell in data:
                    print(f"    {cell}")

    def show_row(self, n):
        n = n - 1  # count from 1 for this function
        if n < self.m_rowcount:
            row = self[n]
            row.show()
        else:
            print("No such row")

    def show_info(self, n):
        n = n - 1  # count from 1 for this function
        if n < self.m_rowcount:
            row = self[n]
            if self.is_row_valid(row):
                info = HazardInfo(row)
                info.show()
            else:
                print("Not a valid hazard data row")
        else:
            print("No such row")

    def get_hazard_info(self, row):
        if type(row) == int:
            return self.get_hazard_info(self[row-1])
        else:
            info = HazardInfo(row-1)
            if info.Valid:
                return info
            else:
                return None

    def find_hazard(self, casnumber):
        for hazard in self.m_hazards:
            if casnumber in hazard.get_cas_numbers():
                return hazard
        return None

    def validate(self, db):
        valid_count = 0
        missing_count = 0
        hazards = self.get_hazard_data()
        for hazard in hazards:
            if hazard.in_database(db):
                valid_count += 1
        print("")
        print(f"{valid_count} of {len(hazards)} hazards were found in the HazardCodes table")
        dbrows = db.queryall("select distinct(CASNumber) from HazardCodes")
        for dbrow in dbrows:
            #print(f"{dbrow}")
            casnum = dbrow['CASNumber'].strip()
            if not self.find_hazard(casnum):
                missing_count += 1
                print(f"CAS # {casnum} is in the database but not in our hazard table")
        print(f"{missing_count} records in the database are not in our hazard table")



    def is_row_valid(self, row):
        if type(row) == int:
            #print(f"is_row_valid: {row}")
            return self.is_row_valid(self[row-1])
        if row.has_hazard_data():
            info = HazardInfo(row)
            return info.Valid
        else:
            return False

    def get_valid_rows(self):
        rows = self.all_rows()
        return [row for row in rows if self.is_row_valid(row)]

    def list_valid_rows(self):
        rows = self.get_valid_rows()
        last_rownum = 0
        for row in rows:
            info = HazardInfo(row)
            rownum = row.m_rownum
            if rownum != last_rownum + 1:
                print("----------")
            last_rownum = rownum
            print(f'Row {rownum+1}: "{info.m_chemical[0:40]}"')





class NewData:
    s_ix_re =  re.compile(r'^\s*\d+-\d+-\d+-\d+\s*$')
    s_cas_re =  re.compile(r'^\s*\d+-\d+-\d+\s*$')
    s_picto_re = re.compile(r'GHS\d\d')

    # format 1
    # ------------------------------------------
    # 4:  INDEX NO: 001-001-00-9
    # 12: CHEMICAL NAME: hydrogen
    # 31: EC NO: 215-605-7
    # 42: CAS #: 1333-74-0
    # 50: CLASS/CAT CODES: Flam.  Gas  1 Press.  Gas
    # 60: HAZARD STMT CODES: H220
    # 69: PICTO: GHS02 GHS04
    #     Dgr
    # 76: HAZARD STMT CODES: H220

    # format 2
    # ------------------------------------------
    # 0:  INDEX NO: 001-001-00-9
    # 7:  CHEMICAL NAME: beryllium
    # 26: EC NO: 231-150-7
    # 36: CAS #: 7440-41-7
    # 45: CLASS/CAT CODES: Acute  Tox.  2  * Acute  Tox.  3  * STOT  RE  1
    #                      Eye  Irrit.  2
    #                      STOT  SE  3
    #                      Skin  Irrit.  2
    #                      Skin  Sens.  1
    # 56: HAZARD STMT CODES: H350i H330 H301
    #                        H372  ** H319 H335 H315 H317
    # 65: PICTO: GHS06 GHS08
    #     Dgr
    # 72: HAZARD STMT CODES: H350i H330 H301
    #                        H372  ** H319 H335 H315 H317



    def __init__(self, excel_row, row_number):
        self.m_rowdata = excel_row
        self.m_rownum = row_number
        
        fmt = self.determine_format()
        self.m_format = fmt
        if fmt == 1:
            self.m_index = self.m_rowdata[4].value
            self.m_chemical = self.value(12)
            self.m_casnumbers = self.value(42)
            self.m_pictostr = self.value(69)
            self.m_pictograms = self.parse_pictograms(self.m_pictostr)
        if fmt == 2:
            self.m_index = self.m_rowdata[0].value
            self.m_chemical = self.value(7)
            self.m_casnumbers = self.value(36)
            self.m_pictostr = self.value(65)
            self.m_pictograms = self.parse_pictograms(self.m_pictostr)

    def show(self):
        print(' ')
        print(f"Row {self.m_rownum} - format {self.m_format}")
        print(f'    Index:      {self.m_index}')
        print(f'    Name:       {self.m_chemical}')
        print(f'    CAS#:       {self.m_casnumbers}')
        print(f'    Pictograms: {" ".join(self.m_pictograms)}')


    def to_json(self, fp):
        picto = " ".join(self.m_pictograms)
        fp.write('\n')
        fp.write('    {\n')
        fp.write(f'        "row": {self.m_rownum},\n')
        fp.write(f'        "format": {self.m_format},\n')
        fp.write(f'        "index":  "{self.m_index}",\n')
        fp.write(f'        "name": "{self.m_chemical}",\n')
        fp.write(f'        "cas":  "{self.m_casnumbers}",\n')
        fp.write(f'        "pictograms": "{picto}"\n')
        fp.write("    },\n")

    def parse_pictograms(self, text):
        result = []
        parts = text.split(' ')
        for part in parts:
            val = part.strip()
            if len(val) > 0:
                result.append(val)
        return result

    def value(self, colnum):
        result = self.m_rowdata[colnum].value
        if result:
            #print(f"value({colnum}) = {result} {type(result)}")
            return str(result).replace('\n', ' ').replace('\r', '').strip()
        else:
            return ''

    def determine_format(self):
        result = 0
        rowdata = self.m_rowdata
        if self.s_ix_re.match(self.value(4)) and self.s_cas_re.match(self.value(42)) and self.s_picto_re.match(self.value(69)):
           result = 1
        if self.s_ix_re.match(self.value(0)) and self.s_cas_re.match(self.value(36)) and self.s_picto_re.match(self.value(65)):
           result = 2
        return result


def get_hazard_codes(casnumber, db):
    return db.queryall(f"select * from hazardcodes where casnumber = '{casnumber}'")

def check_rows(sheet, db):
    new_rows = sheet.get_valid_rows()
    for row in new_rows:
        info = HazardInfo(row)
        codes = get_hazard_codes(info.m_casnumbers, db)
        if len(codes) == 0:
            print(f'Cas # {info.m_casnumbers} is not in the database')

def cells(ws, rownum):
    cells = ws[rownum]
    cellcount = len(cells)
    for i in range(cellcount):
        if cells[i].value:
            print(f"{i}: {cells[i].value}")

def check_formats(ws):
    for rownum in range(1,  ws.max_row):
        rowdata = NewData(ws[rownum], rownum)
        if rowdata.m_format > 0:
            rowdata.show()

def to_json(ws):
    with open("hazards.json", "w") as fp:
        fp.write("[\n")
        for rownum in range(1,  ws.max_row):
            rowdata = NewData(ws[rownum], rownum)
            if rowdata.m_format > 0:
                rowdata.to_json(fp)
        fp.write("]\n")


def show_cmds():
    print(" ")
    print("Commands:")
    print(" ")
    print("help      show this text")
    print("validate  validate the data")
    print("show <n>  show spreadsheet row n (counting from 1)")
    print("next      show the next row")
    print(" ")


def parse_command(cmd):
    result = []
    parts = cmd.split(' ')
    for part in parts:
        part = part.strip()
        if len(part) > 0:
            result.append(part.lower())
    return result

def interactive(sheet, db):
    hazards = sheet.get_hazard_data()
    user = None
    pswd = None
    is_active = True
    cursor = 0
    while is_active:
        cmdline = input(": ")
        parts = parse_command(cmdline)
        if len(parts) == 0:
            continue
        verb = parts[0]
        args = parts[1:]
        if verb == 'help':
            show_cmds()
        if verb == 'validate':
            sheet.validate(db)
        elif verb == 'show':
            cursor = int(args[0])
            row = sheet[cursor-1]
            sheet.show_row(cursor)
            if sheet.is_row_valid(row):
                sheet.show_info(cursor)
            else:
                print("Row is not valid")
        elif verb == 'next':
            cursor += 1
            while cursor <= sheet.m_rowcount and sheet.is_row_valid(cursor) == False:
                cursor += 1
            if cursor < sheet.m_rowcount:
                sheet.show_row(cursor)
                row = sheet[cursor-1]
                sheet.show_info(cursor)
        elif verb == 'find':
            casnumber = args[0]
            match = None
            for h in hazards:
                if casnumber in h.m_casnumbers:
                    match = h
                    break
            if match:
                match.show()
            else:
                print("Not found")
        elif verb == 'quit':
            is_active = False
        elif verb == 'exit':
            is_active = False
            quit()
        else:
            print("I didn't understand that command")


#filename = get_arg('-i', '../Data/new_hazard_data_small.xlsx')
filename = get_arg('-i', '../Data/new_hazard_data.xlsx')

print("Loading workbook " + filename)
wb = load_workbook(filename, read_only=True)
ws = wb.active
print("done")

maxrow = ws.max_row
print(f"Max row number: {maxrow}")
sheet = Sheet(ws)

cmsdb = MySQL('localhost', 'cms', 'cms', 'cms')

if have_arg('-i'):
    interactive(sheet, cmsdb)
